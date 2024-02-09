import openpyxl


def no_of_rows(file_name, sheet_name):
    excel_file = openpyxl.load_workbook(file_name)
    sheet = excel_file[sheet_name]
    return sheet.max_row


def read_data(file_name, sheet_name, row_num, col_num):
    excel_file = openpyxl.load_workbook(file_name)
    sheet = excel_file[sheet_name]
    return sheet.cell(row=row_num, column=col_num).value


def write_data(file_name, sheet_name, row_num, col_num, data):
    excel_file = openpyxl.load_workbook(file_name)
    sheet = excel_file[sheet_name]
    sheet.cell(row=row_num, column=col_num).value = data
    excel_file.save(file_name)
